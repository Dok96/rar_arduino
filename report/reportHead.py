from openpyxl import load_workbook
from config import curr_report
import os
import re  # Для обработки лишних пробелов


def rep_head( rep_time_start, rep_time_stop,  rep_product_len ):
    try:
        # Открываем шаблон Excel
        wb = load_workbook(curr_report)
        ws = wb.active  # Используем активный лист

        # Удаляем лишние пробелы из строковых данных
        def clean_string(value):
            if isinstance(value, str):  # Проверяем, является ли значение строкой
                # Удаляем пробелы в начале и конце строки
                value = value.strip()
                # Заменяем несколько пробелов между словами на один
                value = re.sub(r'\s+', ' ', value)
            return value

        # Очищаем данные перед записью

        rep_time_start = clean_string(rep_time_start)
        rep_time_stop = clean_string(rep_time_stop)

        rep_product_len=clean_string(rep_product_len)

        # Записываем данные в определённые ячейки

        ws["C9"] = f"{rep_time_start}"  # Время начала
        ws["C13"] = f"{rep_product_len}"  #производственная длина
        ws["C11"] = f"{rep_time_stop}"  # Время окончания


        # Сохраняем изменения в тот же файл
        wb.save(curr_report)
        print(f"Отчет успешно сохранен в файл: {curr_report}")

    except FileNotFoundError:
        print(f"Ошибка: Шаблон Excel не найден: {curr_report}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")