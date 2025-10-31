from config import db_number,curr_report
from openpyxl import load_workbook
import struct

# сохраняем данные из плк в текущий отчёт xlsx

def report_alarm(plc, number_of_message):
    try:

        wb = load_workbook(curr_report)
        ws = wb.active  # Используем активный лист

        i=0
        y=0

        # Чтение данных из ПЛК
        for  i in range(number_of_message ):

            y=i * 12
            # читаем сообщения Fault c заси и lump
            read_word_message = plc.db_read(db_number, y, 2) # читаем сразу два байта
            # Преобразование массива байтов в целое число
            status_message = int.from_bytes(read_word_message,byteorder='big')

            print(f'status {status_message}')


            # Читаем длину (4 байта)
            read_len_byte = plc.db_read(db_number, y + 2, 4)  # Смещение: i * 6 + 2 (4 байта для status_lenght)
            if not read_len_byte:
                print(f"Не удалось прочитать status_lenght для сообщения {y + 1}")
                continue
            # Преобразование массива байтов в float
            status_len = struct.unpack('>i', read_len_byte)[0]  # '>f' означает big-endian float
            #status_len = round(status_len,1)

            # читаем сообщения Hour
            read_word_hour = plc.db_read(db_number, y+6, 2)  # читаем сразу два байта
            # Преобразование массива байтов в целое число
            status_hour = int.from_bytes(read_word_hour, byteorder='big')

            # читаем сообщения Minute
            read_word_minute = plc.db_read(db_number, y + 8, 2)  # читаем сразу два байта
            # Преобразование массива байтов в целое число
            status_minute = int.from_bytes(read_word_minute, byteorder='big')

            # читаем сообщения Second
            read_word_second = plc.db_read(db_number, y + 10, 2)  # читаем сразу два байта
            # Преобразование массива байтов в целое число
            status_second = int.from_bytes(read_word_second, byteorder='big')

            # = 1 обнаружен дефект Nec, =2 Lump=8 неисправно или не выбрано, =4 восстановлено
            if  status_message ==256:
                rep_fault  = "обнаружен дефект ВПАДИНА"
            elif status_message ==512:
                rep_fault = "обнаружен дефект УТОЛЩЕНИЕ"
            elif status_message ==2048 :
                rep_fault = "устройство Lump неисправно или не выбрано"
            elif status_message ==1024:
                rep_fault = "устройство Lump восстановлено"

            #ЗАСИ = 1 обнаружен дефект= 8 неисправно или не выбрано, =4 восстановлено
            elif status_message ==1:
                rep_fault = "обнаружен пробой изоляции"
            elif status_message ==4:
                rep_fault = "устройство Spark восстановлено"
            elif status_message ==8:
                rep_fault = "устройство Spark неисправно или не выбрано"
            elif status_message ==0:
                rep_fault = "Event unknow"
            else:
                rep_fault = status_message


            ws[f"A{18 + i}"] = f"{i + 1}" # ячейка А (№)

            ws[f"B{18 + i}"] = f"{rep_fault}" # rep_fault # ячейка Б (Событие)

            ws[f"D{18 + i}"] =status_len # ячейка Д (Длина)

            ws[f"E{18 + i}"] = f'{status_hour:02}:{status_minute:02}:{status_second:02}'

        #Сохраняем изменения в тот же файл
        wb.save(curr_report)
        print(f"Отчет успешно сохранен в файл: {curr_report}")

    except Exception as e:
        print(f"Ошибка: {e}")
        return None